# export.py
import numpy as np
import pandas as pd
import csv
from openpyxl import Workbook


def _compute_kinematics(pendulum):
    hist = pd.DataFrame(pendulum.history)
    L1 = pendulum.p1["length"]
    L2 = pendulum.p2["length"]

    t1, t2 = hist["theta1"].values, hist["theta2"].values
    w1, w2 = hist["omega1"].values, hist["omega2"].values

    x1 = L1 * np.sin(t1)
    y1 = -L1 * np.cos(t1)
    x2 = x1 + L2 * np.sin(t2)
    y2 = y1 - L2 * np.cos(t2)

    vx1 = L1 * np.cos(t1) * w1
    vy1 = L1 * np.sin(t1) * w1
    vx2 = vx1 + L2 * np.cos(t2) * w2
    vy2 = vy1 + L2 * np.sin(t2) * w2

    hist["p1pos_x"], hist["p1pos_y"] = x1, y1
    hist["p2pos_x"], hist["p2pos_y"] = x2, y2
    hist["p1v_x"], hist["p1v_y"] = vx1, vy1
    hist["p2v_x"], hist["p2v_y"] = vx2, vy2

    return hist


def export_to_excel(pendulum, filepath, progress_callback=None):
    hist = _compute_kinematics(pendulum)

    wb = Workbook()
    ws = wb.active

    header1 = ["p1mass", "p1velocity", "p1theta0", "p1omega0",
               "p2mass", "p2velocity", "p2theta0", "p2omega0"]
    for col, text in zip("BCDEFGHI", header1):
        ws[f"{col}1"] = text
    ws["K1"] = "dt"

    ws["A2"] = "Default"
    ws["B2"] = pendulum.p1["mass"]
    ws["C2"] = pendulum.p1["omega0"]
    ws["D2"] = pendulum.p1["theta0"]
    ws["E2"] = pendulum.p1["omega0"]
    ws["F2"] = pendulum.p2["mass"]
    ws["G2"] = pendulum.p2["omega0"]
    ws["H2"] = pendulum.p2["theta0"]
    ws["I2"] = pendulum.p2["omega0"]
    ws["K2"] = pendulum.dt

    header3 = ["p1pos_x", "p1pos_y", "p2pos_x", "p2pos_y",
               "p1v_x", "p1v_y", "p2v_x", "p2v_y"]
    for col, text in zip("BCDEFGHI", header3):
        ws[f"{col}3"] = text

    ws["A4"] = "time"

    cols_order = ["time", "p1pos_x", "p1pos_y", "p2pos_x", "p2pos_y",
                  "p1v_x", "p1v_y", "p2v_x", "p2v_y"]

    rows = list(hist[cols_order].itertuples(index=False))
    total = len(rows)
    report_every = max(1, total // 100)  # 1%씩 갱신

    for i, row in enumerate(rows, start=1):
        row_idx = i + 4  # 5행부터 시작
        for col_letter, value in zip("ABCDEFGHI", row):
            ws[f"{col_letter}{row_idx}"] = value

        if progress_callback and (i % report_every == 0 or i == total):
            progress_callback(int(i / total * 90))  # 90%까지는 행 쓰기, 나머지 10%는 저장(save)

    wb.save(filepath)

    if progress_callback:
        progress_callback(100)


def export_to_csv(pendulum, filepath, progress_callback=None):
    hist = _compute_kinematics(pendulum)
    cols_order = ["time", "theta1", "theta2", "omega1", "omega2",
                  "p1pos_x", "p1pos_y", "p2pos_x", "p2pos_y",
                  "p1v_x", "p1v_y", "p2v_x", "p2v_y"]

    rows = hist[cols_order].values.tolist()
    total = len(rows)
    report_every = max(1, total // 100)

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(cols_order)

        for i, row in enumerate(rows, start=1):
            writer.writerow(row)
            if progress_callback and (i % report_every == 0 or i == total):
                progress_callback(int(i / total * 100))